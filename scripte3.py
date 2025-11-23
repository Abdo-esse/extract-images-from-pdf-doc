import os
import csv
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import io

# --- Configurations ---
EXCEL_FILE = "./Informations  Centre de formation U16 .xlsx"
IMG_DIR = "photos_joueurs"
CSV_FILE = "joueurs_clean.csv"

os.makedirs(IMG_DIR, exist_ok=True)

def safe_name(s):
    """Nettoie les noms pour créer des noms de fichiers valides"""
    return re.sub(r"[^\w]", "_", str(s).strip() if s else "unknown")

def extract_images_from_excel(excel_path, output_dir):
    """Extrait les images embarquées dans un fichier Excel"""
    wb = load_workbook(excel_path)
    ws = wb.active
    
    images_data = []
    
    # Extraire toutes les images de la feuille
    if hasattr(ws, '_images'):
        for idx, img in enumerate(ws._images, start=1):
            try:
                # Récupérer l'image
                image_data = img._data()
                
                # Trouver la position de l'image (ligne)
                row = None
                anchor = img.anchor
                
                # Méthode 1: TwoCellAnchor (ancrage à deux cellules)
                if hasattr(anchor, '_from'):
                    row = anchor._from.row + 1  # +1 car openpyxl commence à 0
                    print(f"    Image {idx}: Ligne {row} (TwoCellAnchor)")
                # Méthode 2: OneCellAnchor
                elif hasattr(anchor, 'row'):
                    row = anchor.row + 1
                    print(f"    Image {idx}: Ligne {row} (OneCellAnchor)")
                # Méthode 3: Essayer d'accéder directement
                else:
                    try:
                        row = int(str(anchor).split('row=')[1].split(',')[0]) + 1
                        print(f"    Image {idx}: Ligne {row} (parsing)")
                    except:
                        print(f"    Image {idx}: Position inconnue - {type(anchor).__name__}")
                
                images_data.append({
                    'data': image_data,
                    'row': row,
                    'index': idx
                })
                
            except Exception as e:
                print(f"  ⚠️ Erreur extraction image {idx}: {e}")
    
    return images_data

def main():
    print(f"📂 Ouverture du fichier Excel: {EXCEL_FILE}")
    
    # Charger le fichier Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    print(f"📊 Feuille active: {ws.title}")
    print(f"📏 Dimensions: {ws.max_row} lignes × {ws.max_column} colonnes")
    
    # Extraire les images
    print(f"\n🖼️ Extraction des images...")
    images_data = extract_images_from_excel(EXCEL_FILE, IMG_DIR)
    print(f"  ℹ️ {len(images_data)} images trouvées")
    
    # Créer un mapping image -> ligne
    image_by_row = {}
    images_without_row = []
    
    for img in images_data:
        if img['row']:
            image_by_row[img['row']] = img
        else:
            images_without_row.append(img)
    
    print(f"  ✓ {len(image_by_row)} images avec position détectée")
    print(f"  ✓ {len(images_without_row)} images sans position")
    
    # Si aucune image n'a de position, on les assigne séquentiellement
    if not image_by_row and images_without_row:
        print(f"  ⚠️ Aucune position détectée, attribution séquentielle...")
        for idx, img in enumerate(images_without_row, start=2):  # start=2 pour ignorer l'en-tête
            image_by_row[idx] = img
            print(f"    Image {img['index']} → Ligne {idx}")
    
    # Lire les données et créer le CSV
    print(f"\n📝 Création du fichier CSV: {CSV_FILE}")
    
    with open(CSV_FILE, "w", newline="", encoding="utf-8") as f_csv:
        writer = csv.writer(f_csv)
        
        # Écrire l'en-tête
        header_row = []
        for cell in ws[1]:
            header_row.append(cell.value if cell.value else "")
        header_row.append("Photo")
        writer.writerow(header_row)
        
        # Parcourir les lignes de données
        saved_images = 0
        rows_with_data = 0
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Vérifier si la ligne contient des données
            if not any(cell for cell in row):
                continue
            
            rows_with_data += 1
            row_data = list(row)
            img_path = ""
            
            # Vérifier si une image correspond à cette ligne
            if row_idx in image_by_row:
                img = image_by_row[row_idx]
                
                # Extraire nom et prénom (colonnes 3 et 4, ajuster selon votre structure)
                try:
                    nom = safe_name(row_data[1]) if len(row_data) > 2 else "inconnu"
                    prenom = safe_name(row_data[2]) if len(row_data) > 3 else "inconnu"
                except:
                    nom = "inconnu"
                    prenom = "inconnu"
                
                # Déterminer l'extension de l'image
                try:
                    image = Image.open(io.BytesIO(img['data']))
                    ext = image.format.lower() if image.format else "png"
                except:
                    ext = "png"
                
                img_path = os.path.join(IMG_DIR, f"{prenom}_{nom}_page{row_idx}.{ext}")
                
                # Sauvegarder l'image
                try:
                    with open(img_path, "wb") as f_img:
                        f_img.write(img['data'])
                    saved_images += 1
                    print(f"  ✅ Image sauvegardée: {os.path.basename(img_path)} → {nom} {prenom}")
                except Exception as e:
                    print(f"  ❌ Erreur sauvegarde image ligne {row_idx}: {e}")
                    img_path = ""
            
            # Écrire la ligne dans le CSV
            writer.writerow(row_data + [img_path])
    
    print(f"\n✅ Traitement terminé!")
    print(f"   📊 {rows_with_data} lignes de données traitées")
    print(f"   🖼️ {saved_images} images sauvegardées")
    print(f"   📄 CSV créé: {CSV_FILE}")
    print(f"   📁 Images dans: {IMG_DIR}/")

if __name__ == "__main__":
    main()